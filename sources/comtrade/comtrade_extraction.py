"""
Extraction des donnees de commerce international pour les Comores
==================================================================
Format de sortie : Excel 2 feuilles (inspire du dataset IATI comoresopendata.org)
  - Feuille "Donnees"     : tableau plat tidy, 1 ligne = 1 flux annuel
                            (annee x flux x section douaniere x pays partenaire)
  - Feuille "Metadonnees" : dictionnaire des colonnes (colonne | source | description)

Source principale :
  WITS/Comtrade (Banque Mondiale) - API SDMX publique, sans cle
  1 requete : reporter=COM, partner=all, product=all
  -> donne section x partenaire x annee pour exports ET imports

Conversions monetaires :
  USD -> EUR : taux annuel moyen BCE (ECB Statistical Data Warehouse, serie EXR)
  EUR -> KMF : taux fixe officiel 1 EUR = 491.968 KMF (franc comorien arrime a l euro)

Pays : Union des Comores (COM / KM)
Derniere mise a jour : 2026-03-23
"""

import requests
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path
from datetime import datetime
import time

# ---------------------------------------------------------
# CONFIG
# ---------------------------------------------------------
COUNTRY_WITS  = "COM"
OUTPUT_DIR    = Path(__file__).parent.parent.parent / "outputs"
OUTPUT_FILE   = OUTPUT_DIR / "flux_commerciaux_comores_exports_imports_par_section.xlsx"
WITS_BASE     = "https://wits.worldbank.org/API/V1/SDMX/V21/datasource/tradestats-trade"
ECB_URL       = "https://data-api.ecb.europa.eu/service/data/EXR/A.USD.EUR.SP00.A?format=jsondata&startPeriod=1990&endPeriod=2025"
ECB_TIMEOUT   = 45
KMF_PER_EUR   = 491.968   # taux fixe officiel, immuable
REQUEST_DELAY = 1.5

# ---------------------------------------------------------
# REFERENTIELS SECTIONS HS
# Un code valide commence par un chiffre (ex: "06-15_Vegetable")
# Les codes UNCTAD comme "Food", "manuf", "UNCTAD-SoP1" sont exclus
# ---------------------------------------------------------
HS_SECTION_LABELS = {
    "01-05_Animal"    : "Produits animaux (ch.01-05)",
    "06-15_Vegetable" : "Produits vegetaux - vanille, girofle (ch.06-15)",
    "16-24_FoodProd"  : "Produits alimentaires transformes (ch.16-24)",
    "25-26_Minerals"  : "Mineraux, pierres (ch.25-26)",
    "27-27_Fuels"     : "Combustibles, hydrocarbures (ch.27)",
    "28-38_Chemicals" : "Produits chimiques - ylang-ylang (ch.28-38)",
    "39-40_PlastiRub" : "Plastiques et caoutchouc (ch.39-40)",
    "41-43_HidesSkin" : "Cuirs, peaux, fourrures (ch.41-43)",
    "44-49_Wood"      : "Bois, papier, edition (ch.44-49)",
    "50-63_TextCloth" : "Textiles et vetements (ch.50-63)",
    "64-67_Footwear"  : "Chaussures, coiffures (ch.64-67)",
    "68-71_StoneGlas" : "Pierres, verre, bijoux (ch.68-71)",
    "72-83_Metals"    : "Metaux et ouvrages en metaux (ch.72-83)",
    "84-85_MachElec"  : "Machines, equipements electriques (ch.84-85)",
    "86-89_Transport" : "Vehicules, bateaux, avions (ch.86-89)",
    "90-99_Miscellan" : "Instruments, optique, divers (ch.90-99)",
}

PARTNER_LABELS = {
    "FRA": "France",           "ARE": "Emirats Arabes Unis", "IND": "Inde",
    "SGP": "Singapour",        "JPN": "Japon",               "TUR": "Turquie",
    "CHN": "Chine",            "ZAF": "Afrique du Sud",      "KEN": "Kenya",
    "MYS": "Malaisie",         "THA": "Thailande",           "PAK": "Pakistan",
    "USA": "Etats-Unis",       "GBR": "Royaume-Uni",         "DEU": "Allemagne",
    "BEL": "Belgique",         "NLD": "Pays-Bas",            "ITA": "Italie",
    "SAU": "Arabie Saoudite",  "MDG": "Madagascar",          "TZA": "Tanzanie",
    "MUS": "Maurice",          "MOZ": "Mozambique",          "EGY": "Egypte",
    "IDN": "Indonesie",        "AUS": "Australie",           "BRA": "Bresil",
    "OMN": "Oman",             "QAT": "Qatar",               "KWT": "Koweit",
    "IRN": "Iran",             "PRT": "Portugal",            "ESP": "Espagne",
    "CHE": "Suisse",           "TUN": "Tunisie",             "MAR": "Maroc",
    "NOR": "Norvege",          "SWE": "Suede",               "CAN": "Canada",
    "PAK": "Pakistan",         "HKG": "Hong Kong",           "MYS": "Malaisie",
    "REU": "La Reunion",       "MYT": "Mayotte",             "DJI": "Djibouti",
    "ETH": "Ethiopie",         "SOM": "Somalie",             "SDN": "Soudan",
    "YEM": "Yemen",            "LKA": "Sri Lanka",           "BGD": "Bangladesh",
    "VNM": "Vietnam",          "PHL": "Philippines",         "TWN": "Taiwan",
    "KOR": "Coree du Sud",     "MEX": "Mexique",             "ARG": "Argentine",
}

# ---------------------------------------------------------
# DICTIONNAIRE DES COLONNES (feuille Metadonnees)
# ---------------------------------------------------------
METADATA = [
    ("annee",                "WITS/Comtrade",
     "Annee de reference du flux commercial"),
    ("flux",                 "WITS/Comtrade",
     "Direction du flux : Export (sortie des Comores) ou Import (entree aux Comores)"),
    ("section_code",         "WITS/Comtrade",
     "Code section douaniere WITS (nomenclature HS). Ex: 06-15_Vegetable pour vanille et girofle"),
    ("section_libelle",      "WITS/Comtrade",
     "Libelle de la section douaniere en francais"),
    ("partenaire_code_iso3", "WITS/Comtrade",
     "Code ISO 3166-1 alpha-3 du pays partenaire (ex: FRA, IND, ARE)"),
    ("partenaire_pays",      "WITS/Comtrade",
     "Nom du pays partenaire en francais"),
    ("valeur_usd",          "WITS/Comtrade",
     "Valeur du flux en dollars USD courants (valeur exacte). Source directe WITS/Comtrade (donnee originale en milliers USD multipliee par 1000)."),
    ("valeur_eur",          "BCE / calcul",
     "Valeur en euros courants (valeur exacte). Calcul : valeur_usd / taux_annuel_usd_eur (BCE, moyenne annuelle). Voir colonne taux_usd_eur."),
    ("valeur_kmf",          "Calcul",
     "Valeur en francs comoriens KMF (valeur exacte). Calcul : valeur_eur x 491.968. Taux fixe officiel inchange : 1 EUR = 491.968 KMF."),
    ("taux_usd_eur",         "BCE - EXR",
     "Taux de change annuel moyen USD/EUR utilise pour la conversion. Source : Banque Centrale Europeenne, serie EXR A.USD.EUR.SP00.A."),
    ("reporter_code",        "WITS/Comtrade",
     "Code ISO3 du pays declarant (toujours COM = Union des Comores)"),
    ("reporter_pays",        "WITS/Comtrade",
     "Nom du pays declarant (toujours Union des Comores)"),
    ("date_extraction",      "Systeme",
     "Date d extraction des donnees depuis l API WITS (YYYY-MM-DD)"),
]


# ---------------------------------------------------------
# ETAPE 1 : Recuperer les taux de change USD/EUR (BCE)
# ---------------------------------------------------------
def fetch_ecb_rates():
    """Retourne un dict {annee(int): taux(float)} ex: {2020: 1.1422}
    Taux = nombre de USD pour 1 EUR (moyenne annuelle BCE)."""
    print("  -> Taux USD/EUR annuels (BCE) ...", end=" ", flush=True)
    try:
        r = requests.get(ECB_URL, timeout=ECB_TIMEOUT)
        r.raise_for_status()
        d = r.json()
        # Structure ECB jsondata: dataSets[0].series["0:0:0:0:0"].observations
        obs = d["dataSets"][0]["series"]["0:0:0:0:0"]["observations"]
        periods = d["structure"]["dimensions"]["observation"][0]["values"]
        rates = {}
        for i, period in enumerate(periods):
            year = int(period["id"])
            val  = obs.get(str(i), [None])[0]
            if val is not None:
                rates[year] = float(val)
        print(f"OK  ({len(rates)} annees, {min(rates)}-{max(rates)})")
        return rates
    except Exception as e:
        print(f"ERREUR: {e} -- utilisation taux de secours")
        # Taux de secours approximatifs si l API BCE est indisponible
        return {
            1999:1.0658, 2000:0.9236, 2001:0.8956, 2002:0.9454, 2003:1.1312,
            2004:1.2438, 2005:1.2441, 2006:1.2556, 2007:1.3705, 2008:1.4726,
            2009:1.3948, 2010:1.3257, 2011:1.3920, 2012:1.2848, 2013:1.3281,
            2014:1.3285, 2015:1.1095, 2016:1.1069, 2017:1.1297, 2018:1.1810,
            2019:1.1195, 2020:1.1422, 2021:1.1827, 2022:1.0530, 2023:1.0813,
            2024:1.0820,
        }


# ---------------------------------------------------------
# ETAPE 2 : Recuperer les donnees WITS (section x partenaire x annee)
# ---------------------------------------------------------
def fetch_wits(url, label):
    print(f"  -> {label} ...", end=" ", flush=True)
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
        root = ET.fromstring(r.text)
        n_obs = len(root.findall(".//{*}Obs"))
        n_ser = len(root.findall(".//{*}Series"))
        print(f"OK  ({n_ser} series, {n_obs} obs)")
        return root
    except Exception as e:
        print(f"ERREUR: {e}")
        return None


def parse_wits(root):
    rows = []
    for series in root.findall(".//{*}Series"):
        s = dict(series.attrib)
        for obs in series.findall(".//{*}Obs"):
            row = dict(obs.attrib)
            row.update(s)
            rows.append(row)
    return rows


def is_hs_section(code):
    """Retourne True uniquement pour les vrais codes sections HS (commencent par un chiffre)."""
    return bool(code) and code[0].isdigit()


def collect_trade(flow, ecb_rates):
    """
    Itere sur chaque section HS : 1 requete WITS par section avec partner=all.
    Evite le 413 (Request Entity Too Large) de l API WITS.
    flow = "X" (export) ou "M" (import)
    """
    indicator  = "XPRT-TRD-VL" if flow == "X" else "MPRT-TRD-VL"
    flux_label = "Export" if flow == "X" else "Import"
    date_ext   = datetime.now().strftime("%Y-%m-%d")
    rows       = []
    sections   = list(HS_SECTION_LABELS.keys())

    for i, section_code in enumerate(sections, 1):
        label = f"{flux_label}s [{i}/{len(sections)}] {section_code}"
        root = fetch_wits(
            f"{WITS_BASE}/reporter/{COUNTRY_WITS}/year/ALL/partner/all/product/{section_code}/indicator/{indicator}",
            label
        )
        time.sleep(REQUEST_DELAY)
        if root is None:
            continue

        for r in parse_wits(root):
            partner  = r.get("PARTNER", "")
            raw_val  = r.get("OBS_VALUE")

            if partner in ("WLD", ""):
                continue
            if raw_val is None:
                continue

            val_usd_thousands = float(raw_val)
            if val_usd_thousands == 0:
                continue

            annee   = int(r.get("TIME_PERIOD", 0))
            val_usd = round(val_usd_thousands * 1000, 2)
            taux    = ecb_rates.get(annee)
            val_eur = round(val_usd / taux, 2) if taux else None
            val_kmf = round(val_eur * KMF_PER_EUR, 0) if val_eur is not None else None

            rows.append({
                "annee"               : annee,
                "flux"                : flux_label,
                "section_code"        : section_code,
                "section_libelle"     : HS_SECTION_LABELS.get(section_code, section_code),
                "partenaire_code_iso3": partner,
                "partenaire_pays"     : PARTNER_LABELS.get(partner, partner),
                "valeur_usd"          : val_usd,
                "valeur_eur"          : val_eur,
                "valeur_kmf"          : val_kmf,
                "taux_usd_eur"        : taux,
                "reporter_code"       : "COM",
                "reporter_pays"       : "Union des Comores",
                "date_extraction"     : date_ext,
            })

    return rows


# ---------------------------------------------------------
# ETAPE 3 : Ecriture Excel (2 feuilles)
# ---------------------------------------------------------
def write_excel(df, output_path):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    col_widths = {
        "annee"               : 8,
        "flux"                : 10,
        "section_code"        : 22,
        "section_libelle"     : 46,
        "partenaire_code_iso3": 20,
        "partenaire_pays"     : 28,
        "valeur_usd"          : 20,
        "valeur_eur"          : 20,
        "valeur_kmf"          : 22,
        "taux_usd_eur"        : 14,
        "reporter_code"       : 16,
        "reporter_pays"       : 24,
        "date_extraction"     : 18,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ---- Feuille 1 : Donnees ----------------------------
        df.to_excel(writer, sheet_name="Donnees", index=False)
        ws = writer.sheets["Donnees"]

        h_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        h_fill  = PatternFill("solid", fgColor="1F3864")
        h_align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font      = h_font
            cell.fill      = h_fill
            cell.alignment = h_align

        for i, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 18)
        ws.freeze_panes = "A2"

        # ---- Feuille 2 : Metadonnees ------------------------
        df_meta = pd.DataFrame(METADATA, columns=["colonne", "source", "description"])
        df_meta.to_excel(writer, sheet_name="Metadonnees", index=False)
        ws_m = writer.sheets["Metadonnees"]

        m_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        m_fill = PatternFill("solid", fgColor="2E75B6")
        for cell in ws_m[1]:
            cell.font      = m_font
            cell.fill      = m_fill
            cell.alignment = Alignment(horizontal="center")

        ws_m.column_dimensions["A"].width = 26
        ws_m.column_dimensions["B"].width = 18
        ws_m.column_dimensions["C"].width = 75
        ws_m.freeze_panes = "A2"
        for row in ws_m.iter_rows(min_row=2, max_col=3):
            row[2].alignment = Alignment(wrap_text=True)
        for i in range(2, len(df_meta) + 2):
            ws_m.row_dimensions[i].height = 42


# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------
def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("\n" + "="*62)
    print("  COMMERCE INTERNATIONAL COMORES - section x partenaire x annee")
    print("  Format : 2 feuilles (Donnees + Metadonnees)")
    print("  Valeurs : USD | EUR (BCE) | KMF (taux fixe 491.968)")
    print("="*62)

    # 1. Taux de change USD/EUR (BCE)
    print("\n[1/3] Taux de change USD/EUR annuels (BCE)")
    ecb_rates = fetch_ecb_rates()

    # 2. Donnees commerciales WITS
    print("\n[2/3] Flux commerciaux WITS (section x partenaire x annee)")
    rows = []
    rows += collect_trade("X", ecb_rates)
    rows += collect_trade("M", ecb_rates)

    # 3. Construction du DataFrame
    print("\n[3/3] Construction et ecriture Excel")
    df = pd.DataFrame(rows)
    if df.empty:
        print("  ERREUR : aucune donnee collectee. Verifiez la connexion API WITS.")
        return
    df = df.sort_values(["flux", "annee", "section_code", "partenaire_code_iso3"]).reset_index(drop=True)

    n_total    = len(df)
    n_export   = int((df["flux"] == "Export").sum())
    n_import   = int((df["flux"] == "Import").sum())
    n_sections = df["section_code"].nunique()
    n_partners = df["partenaire_code_iso3"].nunique()
    years      = sorted(df["annee"].unique())
    yr_range   = f"{min(years)}-{max(years)}" if years else "?"

    print(f"\n{'='*62}")
    print(f"  Total lignes       : {n_total:,}")
    print(f"    Exports          : {n_export:,}")
    print(f"    Imports          : {n_import:,}")
    print(f"  Sections HS        : {n_sections}")
    print(f"  Pays partenaires   : {n_partners}")
    print(f"  Couverture annees  : {yr_range}")
    print(f"  Colonnes           : {len(df.columns)}")
    print(f"{'='*62}\n")

    write_excel(df, OUTPUT_FILE)

    size_kb = OUTPUT_FILE.stat().st_size / 1024
    print(f"  OK : {OUTPUT_FILE.name}  ({size_kb:.0f} KB)")
    print(f"     Feuille Donnees     : {n_total:,} lignes x {len(df.columns)} colonnes")
    print(f"     Feuille Metadonnees : {len(METADATA)} colonnes documentees\n")


if __name__ == "__main__":
    main()
